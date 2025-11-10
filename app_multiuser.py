# app_multiuser.py - Multi-user Flask application with authentication
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from functools import wraps
import sqlite3
from datetime import datetime
import json
import traceback
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Import our modules
import database_multiuser as database
import utils_multiuser as utils
from auth import User, init_auth_db

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'jobfinder-secret-key-change-in-production-xyz123')

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    return User.get(int(user_id))

# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('You need administrator privileges to access this page.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Initialize databases on startup
def init_app():
    """Initialize the application"""
    # Create new multi-user database
    database.init_multiuser_db()
    init_auth_db()
    # Reset scan flags since threads don't persist across restarts
    utils.reset_all_scan_flags()
    print("✅ Application initialized with multi-user support")

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')

        # Special handling for admin username (not an email)
        if email.lower() == 'admin':
            email = 'admin'
        else:
            email = email.lower()

        user = User.verify_password(email, password)
        if user:
            if not user.is_approved:
                flash('Your account is pending approval. Please wait for an administrator to approve your account.', 'info')
                return render_template('login.html')

            login_user(user, remember=True)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Invalid username/email or password.', 'error')

    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Registration page"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        # Validation
        if not name or not email or not password:
            flash('All fields are required.', 'error')
            return render_template('register.html')

        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('register.html')

        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('register.html')

        # Create user
        user_id = User.create_user(email, password, name, is_admin=False, is_approved=False)
        if user_id:
            flash('Registration successful! Your account is pending approval. You will be able to login once an administrator approves your account.', 'success')
            return redirect(url_for('login'))
        else:
            flash('An account with this email already exists.', 'error')

    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    """Logout current user"""
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# Admin routes
@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    """Admin dashboard"""
    users = User.get_all_users()
    return render_template('admin.html', users=users)

@app.route('/admin/approve/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def admin_approve_user(user_id):
    """Approve a user"""
    try:
        User.approve_user(user_id)
        return jsonify({'success': True, 'message': 'User approved successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id):
    """Delete a user"""
    try:
        if user_id == current_user.id:
            return jsonify({'success': False, 'message': 'Cannot delete your own account'})

        User.delete_user(user_id)
        return jsonify({'success': True, 'message': 'User deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/create-user', methods=['POST'])
@login_required
@admin_required
def admin_create_user():
    """Create a new user (admin only)"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        is_admin = data.get('is_admin', False)
        is_approved = data.get('is_approved', True)

        if not name or not email or not password:
            return jsonify({'success': False, 'message': 'All fields are required'})

        user_id = User.create_user(email, password, name, is_admin, is_approved)
        if user_id:
            return jsonify({'success': True, 'message': 'User created successfully'})
        else:
            return jsonify({'success': False, 'message': 'User with this email already exists'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Main application routes (with user context)
@app.route('/')
@login_required
def dashboard():
    """Main dashboard showing approved jobs"""
    try:
        with database.get_conn() as conn:
            # Get approved jobs for current user
            query = """
            SELECT
                a.id as approved_id,
                a.date_approved,
                a.reason,
                a.date_applied,
                a.is_archived,
                d.job_id,
                d.title,
                d.url,
                d.location,
                d.keyword,
                d.description
            FROM approved_jobs a
            JOIN discovered_jobs d ON a.discovered_job_id = d.id
            WHERE a.user_id = ?
                AND (a.is_archived IS NULL OR a.is_archived = FALSE)
                AND (a.is_dismissed IS NULL OR a.is_dismissed = FALSE)
            ORDER BY a.date_approved DESC
            """
            approved_jobs = conn.execute(query, (current_user.id,)).fetchall()

            # Get summary statistics for current user
            stats_query = """
            SELECT
                COUNT(*) as total_discovered,
                (SELECT COUNT(*) FROM approved_jobs WHERE user_id = ? AND (is_archived IS NULL OR is_archived = FALSE) AND (is_dismissed IS NULL OR is_dismissed = FALSE)) as total_approved,
                (SELECT COUNT(*) FROM approved_jobs WHERE user_id = ? AND date_applied IS NOT NULL AND (is_archived IS NULL OR is_archived = FALSE)) as total_applied,
                (SELECT COUNT(*) FROM discovered_jobs WHERE user_id = ? AND analyzed = TRUE) as total_analyzed
            FROM discovered_jobs WHERE user_id = ?
            """
            stats = conn.execute(stats_query, (current_user.id, current_user.id, current_user.id, current_user.id)).fetchone()

        return render_template('dashboard.html',
                             jobs=approved_jobs,
                             stats=stats,
                             scan_status=database.get_scan_status(current_user.id))

    except Exception as e:
        flash(f'Error loading dashboard: {str(e)}', 'error')
        return render_template('dashboard.html', jobs=[], stats=None, scan_status=database.get_scan_status(current_user.id))

@app.route('/api/scan/start', methods=['POST'])
@login_required
def api_start_scan():
    """Start the job scanning process for current user"""
    success, message = utils.start_scan_for_user(current_user.id)
    return jsonify({'success': success, 'message': message})

@app.route('/api/scan/stop', methods=['POST'])
@login_required
def api_stop_scan():
    """Stop the job scanning process for current user"""
    success, message = utils.stop_scan_for_user(current_user.id)
    return jsonify({'success': success, 'message': message})

@app.route('/api/scan/status')
@login_required
def api_scan_status():
    """Get current scan status for current user"""
    return jsonify(database.get_scan_status(current_user.id))

@app.route('/api/job/<int:approved_id>/apply', methods=['POST'])
@login_required
def mark_job_applied(approved_id):
    """Mark a job as applied for current user"""
    try:
        success = database.mark_job_as_applied(current_user.id, approved_id)
        if success:
            return jsonify({'success': True, 'message': 'Job marked as applied'})
        else:
            return jsonify({'success': False, 'message': 'Job was already applied or not found'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/job/<int:approved_id>/delete', methods=['POST'])
@login_required
def delete_approved_job(approved_id):
    """Delete an approved job for current user"""
    try:
        success = database.delete_approved_job(current_user.id, approved_id)
        if success:
            return jsonify({'success': True, 'message': 'Job deleted'})
        else:
            return jsonify({'success': False, 'message': 'Job not found'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/jobs/archive-applied', methods=['POST'])
@login_required
def archive_applied_jobs():
    """Archive all applied jobs for current user"""
    try:
        count = database.archive_all_applied_jobs(current_user.id)
        return jsonify({'success': True, 'message': f'Archived {count} applied jobs'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/jobs/clear-approved', methods=['POST'])
@login_required
def clear_all_approved():
    """Clear all approved jobs for current user"""
    try:
        count = database.clear_all_approved_jobs(current_user.id)
        return jsonify({'success': True, 'message': f'Cleared {count} approved jobs'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/jobs/clear-discovered', methods=['POST'])
@login_required
def clear_all_discovered():
    """Clear all discovered jobs for current user"""
    try:
        count = database.clear_all_discovered_jobs(current_user.id)
        return jsonify({'success': True, 'message': f'Cleared {count} discovered jobs - fresh start ready!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/config')
@login_required
def config_page():
    """Configuration management page for current user"""
    try:
        config = utils.load_user_config(current_user.id)
        project_info = utils.get_project_info()
        presets = utils.get_user_presets(current_user.id)

        return render_template('config.html', config=config, project_info=project_info, presets=presets)
    except Exception as e:
        flash(f'Error loading configuration: {str(e)}', 'error')
        return render_template('config.html', config={}, project_info={}, presets=[])

@app.route('/api/config/save', methods=['POST'])
@login_required
def save_config_api():
    """Save configuration changes for current user"""
    try:
        config_data = request.json

        # Basic validation
        if not config_data:
            return jsonify({'success': False, 'message': 'No configuration data provided'})

        # Save the configuration for current user
        utils.save_user_config(current_user.id, config_data)

        return jsonify({'success': True, 'message': 'Configuration saved successfully'})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saving configuration: {str(e)}'})

# Preset management routes
@app.route('/api/presets/list')
@login_required
def api_list_presets():
    """Get list of available presets for current user"""
    try:
        presets = utils.get_user_presets(current_user.id)
        return jsonify({'success': True, 'presets': presets})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error loading presets: {str(e)}'})

@app.route('/api/presets/save', methods=['POST'])
@login_required
def api_save_preset():
    """Save current configuration as a preset for current user"""
    try:
        data = request.json
        preset_name = data.get('name', '').strip()
        display_name = data.get('display_name', '').strip()
        description = data.get('description', '').strip()

        if not preset_name:
            return jsonify({'success': False, 'message': 'Preset name is required'})

        # Get current configuration
        current_config = utils.load_user_config(current_user.id)

        # Save the preset
        success = utils.save_user_preset(current_user.id, preset_name, current_config, display_name, description)

        if success:
            return jsonify({'success': True, 'message': f'Preset "{display_name or preset_name}" saved successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to save preset'})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saving preset: {str(e)}'})

@app.route('/api/presets/load/<preset_name>')
@login_required
def api_load_preset(preset_name):
    """Load a specific preset configuration for current user"""
    try:
        preset_data = utils.load_user_preset(current_user.id, preset_name)
        if preset_data:
            return jsonify({'success': True, 'preset': preset_data})
        else:
            return jsonify({'success': False, 'message': 'Preset not found'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error loading preset: {str(e)}'})

@app.route('/api/presets/apply/<preset_name>', methods=['POST'])
@login_required
def api_apply_preset(preset_name):
    """Apply a preset as the current configuration for current user"""
    try:
        # Load the preset
        preset_data = utils.load_user_preset(current_user.id, preset_name)
        if not preset_data:
            return jsonify({'success': False, 'message': 'Preset not found'})

        # Apply the preset config
        config = preset_data.get('config', {})
        success = utils.save_user_config(current_user.id, config)

        if success:
            display_name = preset_data.get('display_name', preset_name)
            return jsonify({'success': True, 'message': f'Applied preset "{display_name}" successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to apply preset'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error applying preset: {str(e)}'})

@app.route('/api/presets/delete/<preset_name>', methods=['POST'])
@login_required
def api_delete_preset(preset_name):
    """Delete a preset for current user"""
    try:
        # Get preset info before deleting for display name
        preset_data = utils.load_user_preset(current_user.id, preset_name)
        display_name = preset_data.get('display_name', preset_name) if preset_data else preset_name

        success = utils.delete_user_preset(current_user.id, preset_name)
        if success:
            return jsonify({'success': True, 'message': f'Deleted preset "{display_name}" successfully'})
        else:
            return jsonify({'success': False, 'message': 'Preset not found'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting preset: {str(e)}'})

@app.route('/api/presets/delete-all', methods=['POST'])
@login_required
def api_delete_all_presets():
    """Delete all presets for current user"""
    try:
        # Get all presets first
        presets = utils.get_user_presets(current_user.id)
        deleted_count = 0

        # Delete each preset
        for preset in presets:
            if utils.delete_user_preset(current_user.id, preset['preset_name']):
                deleted_count += 1

        return jsonify({'success': True, 'message': f'Deleted {deleted_count} presets successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting presets: {str(e)}'})

@app.route('/api/presets/create-defaults', methods=['POST'])
@login_required
def api_create_default_presets():
    """Create default presets for current user"""
    try:
        # Create a few default presets
        default_presets = [
            {
                'name': 'entry_level',
                'display_name': 'Entry Level Search',
                'description': 'Optimized for entry-level positions',
                'config': {
                    'search_parameters': {
                        'keywords': ['entry level software engineer', 'junior developer', 'associate engineer'],
                        'locations': ['Remote', 'United States'],
                        'exclusion_keywords': ['senior', 'sr.', 'lead', 'principal', 'staff', 'manager'],
                        'experience_level': 'entry',
                        'job_type': 'full-time',
                        'max_jobs_per_search': 50
                    },
                    'prompts': {
                        'evaluation_prompt': utils.get_default_config()['prompts']['evaluation_prompt']
                    },
                    'api_keys': utils.load_user_config(current_user.id).get('api_keys', {}),
                    'general': utils.get_default_config()['general'],
                    'resume': {'text': ''}
                }
            }
        ]

        created_count = 0
        for preset in default_presets:
            if utils.save_user_preset(current_user.id, preset['name'], preset['config'], preset['display_name'], preset['description']):
                created_count += 1

        return jsonify({'success': True, 'message': f'Created {created_count} default presets'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error creating default presets: {str(e)}'})

@app.route('/job/<int:job_id>')
@login_required
def job_detail(job_id):
    """Job detail page for current user"""
    try:
        with database.get_conn() as conn:
            query = """
            SELECT
                a.id as approved_id,
                a.date_approved,
                a.reason,
                a.date_applied,
                a.is_archived,
                d.job_id,
                d.title,
                d.url,
                d.location,
                d.keyword,
                d.description,
                d.date_discovered
            FROM approved_jobs a
            JOIN discovered_jobs d ON a.discovered_job_id = d.id
            WHERE d.job_id = ? AND a.user_id = ?
            """
            job = conn.execute(query, (job_id, current_user.id)).fetchone()

            if not job:
                flash('Job not found', 'error')
                return redirect(url_for('dashboard'))

            return render_template('job_detail.html', job=job)

    except Exception as e:
        flash(f'Error loading job details: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/statistics')
@login_required
def statistics_page():
    """Statistics and analytics page for current user"""
    try:
        statistics = database.get_job_statistics(current_user.id)
        return render_template('statistics.html',
                             statistics=statistics,
                             scan_status=database.get_scan_status(current_user.id))

    except Exception as e:
        flash(f'Error loading statistics: {str(e)}', 'error')
        empty_stats = {
            'basic': {'total_discovered': 0, 'total_analyzed': 0, 'total_with_details': 0},
            'approved': {'total_approved': 0, 'total_applied': 0, 'total_archived': 0},
            'by_location': [],
            'by_keyword': [],
            'recent_activity': []
        }
        return render_template('statistics.html', statistics=empty_stats, scan_status=database.get_scan_status(current_user.id))

@app.route('/applied')
@login_required
def archived_page():
    """Applied jobs page for current user"""
    try:
        archived_jobs = database.get_archived_jobs(current_user.id)
        return render_template('archived.html',
                             jobs=archived_jobs,
                             scan_status=database.get_scan_status(current_user.id))
    except Exception as e:
        flash(f'Error loading applied jobs: {str(e)}', 'error')
        return render_template('archived.html', jobs=[], scan_status=database.get_scan_status(current_user.id))

@app.route('/api/applied/export')
@login_required
def export_applied_jobs():
    """Export applied jobs to Excel spreadsheet"""
    try:
        # Get applied jobs for current user
        archived_jobs = database.get_archived_jobs(current_user.id)

        if not archived_jobs:
            flash('No applied jobs to export', 'info')
            return redirect(url_for('archived_page'))

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Applied Jobs"

        # Define styles
        header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Define headers
        headers = [
            "Job Title",
            "Company/Position",
            "Location",
            "Keyword Match",
            "LinkedIn URL",
            "Date Applied",
            "Date Approved",
            "AI Match Reasoning",
            "Job Description"
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Set column widths
        column_widths = {
            'A': 30,  # Job Title
            'B': 25,  # Company
            'C': 20,  # Location
            'D': 20,  # Keyword
            'E': 50,  # LinkedIn URL
            'F': 15,  # Date Applied
            'G': 15,  # Date Approved
            'H': 60,  # AI Reasoning
            'I': 70   # Description
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Write data rows
        for row_num, job in enumerate(archived_jobs, 2):
            # Format dates
            date_applied = ''
            if job['date_applied']:
                try:
                    dt = datetime.fromisoformat(job['date_applied'].replace('Z', '+00:00'))
                    date_applied = dt.strftime('%Y-%m-%d %H:%M')
                except:
                    date_applied = str(job['date_applied'])

            date_approved = ''
            if job['date_approved']:
                try:
                    dt = datetime.fromisoformat(job['date_approved'].replace('Z', '+00:00'))
                    date_approved = dt.strftime('%Y-%m-%d %H:%M')
                except:
                    date_approved = str(job['date_approved'])

            # Prepare row data
            row_data = [
                job.get('title') or 'Untitled Position',
                job.get('title') or 'N/A',  # Company - using title as fallback
                job.get('location') or 'N/A',
                job.get('keyword') or 'N/A',
                job.get('url') or 'N/A',
                date_applied,
                date_approved,
                job.get('reason') or 'No reasoning provided',
                job.get('description') or 'No description available'
            ]

            # Write row
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = cell_alignment
                cell.border = border

                # Make URL a hyperlink
                if col_num == 5 and value and value != 'N/A':
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")

        # Freeze first row
        ws.freeze_panes = "A2"

        # Set row height for header
        ws.row_dimensions[1].height = 30

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'applied_jobs_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error exporting applied jobs: {str(e)}")
        print(traceback.format_exc())
        flash(f'Error exporting jobs: {str(e)}', 'error')
        return redirect(url_for('archived_page'))

@app.route('/logs')
@login_required
def logs_page():
    """Logs and system information page for current user"""
    try:
        with database.get_conn() as conn:
            recent_discovered = conn.execute("""
                SELECT job_id, title, url, location, date_discovered, analyzed
                FROM discovered_jobs
                WHERE user_id = ?
                ORDER BY date_discovered DESC
                LIMIT 50
            """, (current_user.id,)).fetchall()

            recent_approved = conn.execute("""
                SELECT
                    a.date_approved,
                    a.reason,
                    d.job_id,
                    d.title,
                    d.url
                FROM approved_jobs a
                JOIN discovered_jobs d ON a.discovered_job_id = d.id
                WHERE a.user_id = ?
                ORDER BY a.date_approved DESC
                LIMIT 20
            """, (current_user.id,)).fetchall()

        return render_template('logs.html',
                             recent_discovered=recent_discovered,
                             recent_approved=recent_approved,
                             scan_status=database.get_scan_status(current_user.id),
                             project_info=utils.get_project_info())

    except Exception as e:
        flash(f'Error loading logs: {str(e)}', 'error')
        return render_template('logs.html', recent_discovered=[], recent_approved=[])

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

# Template filters for better formatting
@app.template_filter('datetime_format')
def datetime_format(value):
    """Format datetime for display"""
    if value is None:
        return ''
    if isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%dT%H:%M:%SZ')
        except:
            return value
    return value

@app.template_filter('truncate_text')
def truncate_text(text, length=100):
    """Truncate text to specified length"""
    if not text:
        return ''
    return text[:length] + '...' if len(text) > length else text

if __name__ == '__main__':
    init_app()
    app.run(debug=True, host='0.0.0.0', port=8734)